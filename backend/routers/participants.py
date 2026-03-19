from datetime import datetime, timezone
import asyncio
from dataclasses import dataclass
import logging
import random
import time
from typing import Any, Optional
from io import BytesIO
from pathlib import Path

from fastapi import APIRouter, BackgroundTasks, HTTPException, Header
from fastapi.concurrency import run_in_threadpool
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from jose import JWTError, jwt

from database import supabase
from models import RegisterRequest, UpdateStatusRequest
from services.url_validator import validate_url
from services.email_service import audit_and_notify
from config import (
    SECRET_KEY,
    ALGORITHM,
    REGISTER_DB_TIMEOUT_SECONDS,
    REGISTER_DB_MAX_RETRIES,
    REGISTER_DB_RETRY_BASE_DELAY,
    REGISTER_DB_RETRY_JITTER,
    REGISTER_DB_MAX_CONCURRENCY,
    REGISTER_DB_QUEUE_WAIT_SECONDS,
    REGISTER_ASYNC_QUEUE_ENABLED,
    REGISTER_ASYNC_QUEUE_MAXSIZE,
    REGISTER_ASYNC_QUEUE_WORKERS,
    REGISTER_ASYNC_QUEUE_RETRY_LIMIT,
    REGISTRATION_CLOSE_AT,
    REGISTRATION_FORCE_CLOSED,
    REGISTRATION_CLOSED_FLAG_FILE,
)

router = APIRouter()
logger = logging.getLogger("register_service")
_REGISTER_WRITE_SEMAPHORE = asyncio.Semaphore(max(1, REGISTER_DB_MAX_CONCURRENCY))
_REGISTER_FALLBACK_QUEUE: Optional[asyncio.Queue] = None
_REGISTER_QUEUE_WORKERS: list[asyncio.Task] = []


def _parse_close_at() -> Optional[datetime]:
    try:
        parsed = datetime.fromisoformat(REGISTRATION_CLOSE_AT)
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=timezone.utc)
        return parsed.astimezone(timezone.utc)
    except Exception:
        logger.warning("Invalid REGISTRATION_CLOSE_AT: %s", REGISTRATION_CLOSE_AT)
        return None


_REGISTRATION_CLOSE_AT_UTC = _parse_close_at()


def _is_registration_closed() -> bool:
    if REGISTRATION_FORCE_CLOSED:
        return True

    flag_path = Path(REGISTRATION_CLOSED_FLAG_FILE)
    if flag_path.exists():
        return True

    if _REGISTRATION_CLOSE_AT_UTC is None:
        return False

    return datetime.now(timezone.utc) >= _REGISTRATION_CLOSE_AT_UTC


def _registration_status_payload() -> dict[str, object]:
    return {
        "closed": _is_registration_closed(),
        "close_at": REGISTRATION_CLOSE_AT,
        "message": "报名已截止" if _is_registration_closed() else "报名进行中",
    }


@dataclass
class RegisterQueueJob:
    table_name: str
    data: dict[str, object]
    endpoint: str
    attempt: int = 0


def _queue_enabled() -> bool:
    return REGISTER_ASYNC_QUEUE_ENABLED


async def _start_register_queue_workers() -> None:
    global _REGISTER_FALLBACK_QUEUE
    if not _queue_enabled() or _REGISTER_QUEUE_WORKERS:
        return

    _REGISTER_FALLBACK_QUEUE = asyncio.Queue(
        maxsize=max(1, REGISTER_ASYNC_QUEUE_MAXSIZE)
    )

    async def _worker(worker_id: int):
        assert _REGISTER_FALLBACK_QUEUE is not None
        while True:
            job: RegisterQueueJob = await _REGISTER_FALLBACK_QUEUE.get()
            try:
                await _insert_with_resilience(
                    table_name=job.table_name,
                    data=job.data,
                    endpoint=f"{job.endpoint}_queued",
                )
                logger.info(
                    "register_queue_job_ok worker=%d endpoint=%s table=%s attempt=%d",
                    worker_id,
                    job.endpoint,
                    job.table_name,
                    job.attempt,
                )
            except HTTPException as exc:
                retryable = exc.status_code in (429, 503)
                if retryable and job.attempt < REGISTER_ASYNC_QUEUE_RETRY_LIMIT:
                    job.attempt += 1
                    delay = min(
                        2.0, REGISTER_DB_RETRY_BASE_DELAY * (2 ** (job.attempt - 1))
                    )
                    delay += random.uniform(0.0, REGISTER_DB_RETRY_JITTER)
                    await asyncio.sleep(delay)
                    await _REGISTER_FALLBACK_QUEUE.put(job)
                    logger.warning(
                        "register_queue_job_retry worker=%d endpoint=%s table=%s attempt=%d status=%d",
                        worker_id,
                        job.endpoint,
                        job.table_name,
                        job.attempt,
                        exc.status_code,
                    )
                else:
                    logger.error(
                        "register_queue_job_drop worker=%d endpoint=%s table=%s attempt=%d status=%d",
                        worker_id,
                        job.endpoint,
                        job.table_name,
                        job.attempt,
                        exc.status_code,
                    )
            except Exception as exc:
                logger.error(
                    "register_queue_job_exception worker=%d endpoint=%s table=%s error=%s",
                    worker_id,
                    job.endpoint,
                    job.table_name,
                    str(exc),
                )
            finally:
                _REGISTER_FALLBACK_QUEUE.task_done()

    for idx in range(max(1, REGISTER_ASYNC_QUEUE_WORKERS)):
        _REGISTER_QUEUE_WORKERS.append(asyncio.create_task(_worker(idx + 1)))

    logger.info(
        "register_queue_started enabled=%s workers=%d maxsize=%d",
        _queue_enabled(),
        len(_REGISTER_QUEUE_WORKERS),
        REGISTER_ASYNC_QUEUE_MAXSIZE,
    )


async def _stop_register_queue_workers() -> None:
    if not _REGISTER_QUEUE_WORKERS:
        return
    for task in _REGISTER_QUEUE_WORKERS:
        task.cancel()
    await asyncio.gather(*_REGISTER_QUEUE_WORKERS, return_exceptions=True)
    _REGISTER_QUEUE_WORKERS.clear()
    logger.info("register_queue_stopped")


async def initialize_participant_services() -> None:
    await _start_register_queue_workers()


async def shutdown_participant_services() -> None:
    await _stop_register_queue_workers()


async def _enqueue_fallback_job(
    *,
    table_name: str,
    data: dict[str, object],
    endpoint: str,
) -> int:
    if not _queue_enabled() or _REGISTER_FALLBACK_QUEUE is None:
        raise HTTPException(
            status_code=429,
            detail="报名高峰中，请稍后重试（建议30秒后重试）",
        )
    if _REGISTER_FALLBACK_QUEUE.full():
        logger.warning(
            "register_queue_full endpoint=%s table=%s size=%d",
            endpoint,
            table_name,
            _REGISTER_FALLBACK_QUEUE.qsize(),
        )
        raise HTTPException(
            status_code=429,
            detail="报名请求过多，请稍后重试（建议30秒后重试）",
        )

    await _REGISTER_FALLBACK_QUEUE.put(
        RegisterQueueJob(table_name=table_name, data=data, endpoint=endpoint)
    )
    size = _REGISTER_FALLBACK_QUEUE.qsize()
    logger.info(
        "register_queue_enqueue endpoint=%s table=%s queue_size=%d",
        endpoint,
        table_name,
        size,
    )
    return size


def _is_retryable_db_error(exc: Exception) -> bool:
    text = str(exc).lower()
    markers = [
        "timeout",
        "timed out",
        "too many requests",
        "connection reset",
        "connection aborted",
        "connection error",
        "connection pool",
        "code': '429'",
        'code": "429"',
        " 429",
        " 502",
        " 503",
        " 504",
        "server closed",
        "temporarily",
    ]
    return any(m in text for m in markers)


async def _insert_with_resilience(
    *,
    table_name: str,
    data: dict[str, object],
    endpoint: str,
) -> Any:
    start_wait = time.perf_counter()
    try:
        await asyncio.wait_for(
            _REGISTER_WRITE_SEMAPHORE.acquire(),
            timeout=REGISTER_DB_QUEUE_WAIT_SECONDS,
        )
    except asyncio.TimeoutError:
        logger.warning(
            "register_db_queue_timeout endpoint=%s table=%s wait_timeout=%.2f inflight_limit=%d",
            endpoint,
            table_name,
            REGISTER_DB_QUEUE_WAIT_SECONDS,
            REGISTER_DB_MAX_CONCURRENCY,
        )
        raise HTTPException(
            status_code=429,
            detail="报名高峰中，请稍后重试（建议30秒后重试）",
        )

    wait_ms = (time.perf_counter() - start_wait) * 1000
    try:
        last_error: Optional[Exception] = None
        for attempt in range(1, REGISTER_DB_MAX_RETRIES + 2):
            t0 = time.perf_counter()
            try:
                result = await asyncio.wait_for(
                    run_in_threadpool(
                        lambda: supabase.table(table_name).insert(data).execute()
                    ),
                    timeout=REGISTER_DB_TIMEOUT_SECONDS,
                )
                elapsed_ms = (time.perf_counter() - t0) * 1000
                logger.info(
                    "register_db_ok endpoint=%s table=%s attempt=%d wait_ms=%.2f db_ms=%.2f",
                    endpoint,
                    table_name,
                    attempt,
                    wait_ms,
                    elapsed_ms,
                )
                return result
            except asyncio.TimeoutError as exc:
                last_error = exc
                retryable = True
                reason = "timeout"
            except Exception as exc:
                last_error = exc
                retryable = _is_retryable_db_error(exc)
                reason = "retryable" if retryable else "non_retryable"

            logger.warning(
                "register_db_fail endpoint=%s table=%s attempt=%d reason=%s error=%s",
                endpoint,
                table_name,
                attempt,
                reason,
                str(last_error),
            )

            if not retryable or attempt >= REGISTER_DB_MAX_RETRIES + 1:
                raise HTTPException(
                    status_code=503,
                    detail="服务繁忙，请稍后重试",
                )

            backoff = REGISTER_DB_RETRY_BASE_DELAY * (2 ** (attempt - 1))
            jitter = random.uniform(0.0, REGISTER_DB_RETRY_JITTER)
            await asyncio.sleep(backoff + jitter)

        raise HTTPException(status_code=503, detail="服务繁忙，请稍后重试")
    finally:
        _REGISTER_WRITE_SEMAPHORE.release()


class CheckUrlRequest(BaseModel):
    url: str


@router.get("/api/participants/registration-status")
async def get_registration_status():
    return _registration_status_payload()


@router.post("/api/participants/check-url")
async def check_url_accessibility(body: CheckUrlRequest):
    """检查 URL 格式（仅前端格式校验，不做可访问性检查）"""
    try:
        if _is_registration_closed():
            return {
                "accessible": False,
                "error": "报名已截止",
                "registration_closed": True,
            }

        # 仅做前端格式校验
        validation_error = validate_url(body.url, "url")
        if validation_error and validation_error.get("level") == "error":
            return {"accessible": False, "error": validation_error.get("message")}

        # 直接返回成功，不做后端可访问性检查
        return {"accessible": True, "message": "链接格式正确"}
    except Exception as e:
        return {"accessible": False, "error": f"检查失败：{str(e)}"}


@router.post("/api/participants/register")
async def register_participant(
    body: RegisterRequest, background_tasks: BackgroundTasks
):
    if _is_registration_closed():
        raise HTTPException(status_code=403, detail="报名已截止")

    url_issues = []
    for field, value in [
        ("pdfUrl", body.pdfUrl),
        ("posterUrl", body.posterUrl),
        ("videoUrl", body.videoUrl),
        ("repoUrl", body.repoUrl),
        ("demoUrl", body.demoUrl),
    ]:
        issue = validate_url(value, field)
        if issue:
            url_issues.append(issue)

    errors = [i for i in url_issues if i["level"] == "error"]
    if errors:
        raise HTTPException(
            status_code=422,
            detail={
                "url_errors": errors,
                "message": "提交的链接存在问题，请修正后重新提交",
            },
        )

    data: dict[str, object] = {
        "full_name": body.fullName,
        "email": body.email,
        "organization": body.organization,
        "github": body.phone,
        "track": body.track or None,
        "project_title": body.projectTitle,
        "project_description": body.projectDescription,
        "demo_url": body.demoUrl or None,
        "repo_url": body.repoUrl or None,
        "pdf_url": body.pdfUrl or None,
        "video_url": body.videoUrl or None,
        "poster_url": body.posterUrl or None,
        "status": "pending",
    }

    try:
        result = await _insert_with_resilience(
            table_name="participants",
            data=data,
            endpoint="register",
        )

        background_tasks.add_task(
            audit_and_notify,
            name=body.fullName,
            email=body.email,
            project_title=body.projectTitle,
            url_fields={
                "pdfUrl": body.pdfUrl,
                "posterUrl": body.posterUrl,
                "videoUrl": body.videoUrl,
                "repoUrl": body.repoUrl,
                "demoUrl": body.demoUrl,
            },
        )

        warnings = [i for i in url_issues if i["level"] == "warning"]
        return {
            "message": "Registration successful",
            "data": result.data,
            "url_warnings": warnings if warnings else None,
        }

    except HTTPException as exc:
        if exc.status_code == 429:
            queue_size = await _enqueue_fallback_job(
                table_name="participants",
                data=data,
                endpoint="register",
            )
            warnings = [i for i in url_issues if i["level"] == "warning"]
            return {
                "message": "Registration accepted and queued",
                "queued": True,
                "queue_size": queue_size,
                "url_warnings": warnings if warnings else None,
            }
        raise

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/api/judges/participants")
async def get_participants(status: Optional[str] = None):
    try:
        query = supabase.table("participants").select("*")
        if status:
            query = query.eq("status", status)
        result = query.execute()
        return {"data": result.data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/api/judges/participants/stats/tracks")
async def get_track_stats(status: Optional[str] = None):
    """获取赛道统计，可选按状态过滤"""
    try:
        query = supabase.table("participants").select("track")
        if status:
            query = query.eq("status", status)
        result = query.execute()

        counts = {}
        for p in result.data:
            track = p.get("track") or "未分类"
            counts[track] = counts.get(track, 0) + 1
        total = sum(counts.values())
        return {
            "data": [{"track": t, "count": c} for t, c in sorted(counts.items())],
            "total": total,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.patch("/api/judges/participants/{participant_id}/status")
async def update_participant_status(participant_id: int, body: UpdateStatusRequest):
    """更新参赛者状态（初筛阶段）"""
    try:
        valid_statuses = ["pending", "reviewing", "scored", "rejected"]
        if body.status not in valid_statuses:
            raise HTTPException(status_code=400, detail="Invalid status")

        update_data: dict[str, object] = {
            "status": body.status,
            "updated_at": datetime.utcnow().isoformat(),
        }

        # 更新材料完整性字段
        if body.materials_complete is not None:
            update_data["materials_complete"] = body.materials_complete

        result = (
            supabase.table("participants")
            .update(update_data)
            .eq("id", participant_id)
            .execute()
        )

        if not result.data:
            raise HTTPException(status_code=404, detail="Participant not found")

        if body.comments and body.status == "rejected":
            supabase.table("scores").insert(
                {
                    "participant_id": participant_id,
                    "judge_id": None,
                    "innovation_score": 0,
                    "technical_score": 0,
                    "market_score": 0,
                    "demo_score": 0,
                    "weighted_score": 0,
                    "comments": f"初筛不通过: {body.comments}",
                    "created_at": datetime.utcnow().isoformat(),
                }
            ).execute()

        return {"message": "Status updated successfully", "data": result.data}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/api/judges/participants/{participant_id}/next")
async def get_next_participant(participant_id: int, status: Optional[str] = None):
    """获取当前参赛者之后的下一个参赛者"""
    try:
        query = supabase.table("participants").select("id, status").order("id")
        if status:
            query = query.eq("status", status)
        result = query.execute()
        all_ids = [p["id"] for p in result.data]

        if participant_id in all_ids:
            idx = all_ids.index(participant_id)
            if idx + 1 < len(all_ids):
                return {"data": {"next_id": all_ids[idx + 1]}}
        return {"data": {"next_id": None}}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/api/judges/participants/{participant_id}")
async def get_participant(participant_id: int):
    try:
        result = (
            supabase.table("participants")
            .select("*")
            .eq("id", participant_id)
            .execute()
        )
        if not result.data:
            raise HTTPException(status_code=404, detail="Participant not found")
        return {"data": result.data[0]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/api/judges/participants/{participant_id}")
async def delete_participant(participant_id: int):
    """删除参赛者及其相关评分记录"""
    try:
        supabase.table("scores").delete().eq("participant_id", participant_id).execute()
        result = (
            supabase.table("participants").delete().eq("id", participant_id).execute()
        )

        if not result.data:
            raise HTTPException(status_code=404, detail="Participant not found")

        return {"message": "Participant deleted successfully", "data": result.data}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/api/judges/participants/export/excel")
async def export_participants_excel(authorization: Optional[str] = Header(None)):
    """导出参赛者评审情况为Excel"""
    # 验证 token
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="未授权")

    token = authorization.replace("Bearer ", "")
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        if payload.get("role") != "judge":
            raise HTTPException(status_code=403, detail="权限不足")
    except JWTError:
        raise HTTPException(status_code=401, detail="无效的令牌")

    try:
        # 获取所有参赛者数据
        result = supabase.table("participants").select("*").order("id").execute()
        participants = result.data

        # 创建Excel工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "参赛者评审情况"

        # 设置表头（移除姓名和机构以保证评审公平性）
        headers = [
            "ID",
            "赛道",
            "项目标题",
            "项目描述",
            "Demo链接",
            "代码仓库",
            "PDF文档",
            "视频链接",
            "海报链接",
            "材料是否齐全",
            "是否通过",
            "备注信息",
            "提交时间",
        ]

        # 写入表头并设置样式
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 写入数据
        track_map = {
            "academic": "学术龙虾",
            "productivity": "生产力龙虾",
            "life": "生活龙虾",
        }

        status_map = {
            "pending": "待评审",
            "reviewing": "评审中",
            "scored": "已评分",
            "rejected": "已拒绝",
        }

        for row_num, p in enumerate(participants, 2):
            # 材料是否齐全
            materials_status = ""
            if p.get("materials_complete") is True:
                materials_status = "齐全"
            elif p.get("materials_complete") is False:
                materials_status = "不齐全"
            else:
                materials_status = "待审核"

            # 是否通过
            pass_status = ""
            if p.get("status") == "reviewing":
                pass_status = "通过"
            elif p.get("status") == "rejected":
                pass_status = "不通过"
            else:
                pass_status = "待定"

            # 备注信息（从scores表获取最近一条非空备注）
            comments = ""
            try:
                score_result = (
                    supabase.table("scores")
                    .select("comments, created_at")
                    .eq("participant_id", p["id"])
                    .order("created_at", desc=True)
                    .execute()
                )
                if score_result.data:
                    for score in score_result.data:
                        text = (score.get("comments") or "").strip()
                        if text:
                            comments = text
                            break
            except Exception:
                pass

            row_data = [
                p.get("id"),
                track_map.get(p.get("track"), p.get("track", "")),
                p.get("project_title"),
                p.get("project_description"),
                p.get("demo_url", ""),
                p.get("repo_url", ""),
                p.get("pdf_url", ""),
                p.get("video_url", ""),
                p.get("poster_url", ""),
                materials_status,
                pass_status,
                comments,
                p.get("created_at", ""),
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        # 调整列宽
        column_widths = [8, 15, 30, 40, 35, 35, 35, 35, 35, 15, 12, 40, 20]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col_num)
            ].width = width

        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # 返回文件
        filename = f"participants_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
