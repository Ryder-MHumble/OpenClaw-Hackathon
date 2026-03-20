from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from database import supabase


def main() -> None:
    output_dir = Path("../exports").resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = (
        output_dir
        / f"projects_by_track_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

    track_map = [
        ("academic", "学术赛道"),
        ("productivity", "生产力赛道"),
        ("life", "生活赛道"),
    ]
    status_map = {
        "pending": "待初筛",
        "reviewing": "初筛通过",
        "rejected": "已拒绝",
        "scored": "已评分",
    }
    headers = [
        "ID",
        "项目名称",
        "团队/机构",
        "联系人",
        "邮箱",
        "当前状态",
        "赛道",
        "提交时间",
        "计划书URL",
        "演示视频URL",
        "海报URL",
        "Demo URL",
        "Repo URL",
        "材料完整性",
        "备注",
        "评价",
    ]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for track_id, sheet_name in track_map:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(headers)

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="1F2937", end_color="1F2937", fill_type="solid"
            )
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

        rows = (
            supabase.table("participants")
            .select("*")
            .eq("track", track_id)
            .order("created_at", desc=False)
            .execute()
            .data
        )

        for p in rows:
            ws.append(
                [
                    p.get("id"),
                    p.get("project_title") or "",
                    p.get("organization") or "",
                    p.get("full_name") or "",
                    p.get("email") or "",
                    status_map.get(p.get("status"), p.get("status") or ""),
                    sheet_name,
                    p.get("created_at") or "",
                    p.get("pdf_url") or "",
                    p.get("video_url") or "",
                    p.get("poster_url") or "",
                    p.get("demo_url") or "",
                    p.get("repo_url") or "",
                    "不完整"
                    if p.get("materials_complete") is False
                    else ("完整" if p.get("materials_complete") is True else ""),
                    p.get("comments") or "",
                    "",
                ]
            )

        widths = [8, 36, 24, 14, 24, 12, 12, 22, 32, 32, 32, 28, 28, 10, 22, 20]
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

        for row in ws.iter_rows(
            min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)
        ):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        ws.freeze_panes = "A2"

    wb.save(output_path)
    print(output_path)


if __name__ == "__main__":
    main()
