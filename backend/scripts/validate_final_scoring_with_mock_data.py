#!/usr/bin/env python3
"""Generate realistic judge scores and validate final scoring logic.

Outputs:
1) database/exports/judge_scores_realistic_mock.csv
2) database/exports/team_score_validation_report.csv
3) database/exports/final_scoring_full_pipeline_audit.xlsx
"""

from __future__ import annotations

import csv
import random
from datetime import datetime, timezone
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import cast

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


WEIGHTS = {
    "innovation": 0.2,
    "technical": 0.2,
    "market": 0.5,
    "demo": 0.1,
}


def weighted_score(
    innovation: float, technical: float, market: float, demo: float
) -> float:
    return (
        innovation * WEIGHTS["innovation"]
        + technical * WEIGHTS["technical"]
        + market * WEIGHTS["market"]
        + demo * WEIGHTS["demo"]
    )


def round_to_half(value: float) -> float:
    return round(value * 2) / 2


def js_like_round_1dp(value: float) -> float:
    """Approximate JS Number.toFixed(1) half-up behavior."""
    return float(Decimal(str(value)).quantize(Decimal("0.1"), rounding=ROUND_HALF_UP))


def round_half_up(value: float, digits: int) -> float:
    quant = Decimal("1").scaleb(-digits)
    return float(Decimal(str(value)).quantize(quant, rounding=ROUND_HALF_UP))


def apply_header_style(ws, headers: list[str]) -> None:
    fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def autosize_columns(ws, min_width: int = 12, max_width: int = 48) -> None:
    for col_cells in ws.columns:
        col_idx = col_cells[0].column
        length = 0
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            length = max(length, len(value))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(
            min_width, min(max_width, length + 2)
        )


def export_pipeline_xlsx(
    *,
    raw_rows: list[dict],
    report_rows: list[dict],
    formula_mismatch_count: int,
    xlsx_path: Path,
) -> None:
    wb = Workbook()

    ws_overview = cast(Worksheet, wb.active)
    ws_overview.title = "00_Overview"
    overview_headers = ["Item", "Value"]
    apply_header_style(ws_overview, overview_headers)
    overview_rows = [
        ["Generated At (UTC)", datetime.now(timezone.utc).isoformat()],
        ["Audit Scope", "Mock final scoring full-pipeline audit"],
        ["Backend Formula", "innovation*0.2 + technical*0.2 + market*0.5 + demo*0.1"],
        ["Frontend Preview Formula", "Same formula, displayed with toFixed(1)"],
        ["Raw Score Rows", len(raw_rows)],
        ["Team Count", len(report_rows)],
        ["Formula Mismatch Count", formula_mismatch_count],
        [
            "Teams With 1dp Display Collision",
            sum(1 for r in report_rows if r["display_collision_1dp"]),
        ],
        ["Weight Innovation", WEIGHTS["innovation"]],
        ["Weight Technical", WEIGHTS["technical"]],
        ["Weight Market", WEIGHTS["market"]],
        ["Weight Demo", WEIGHTS["demo"]],
    ]
    for r, (k, v) in enumerate(overview_rows, 2):
        ws_overview.cell(row=r, column=1, value=k)
        ws_overview.cell(row=r, column=2, value=v)
    ws_overview.freeze_panes = "A2"
    autosize_columns(ws_overview)

    ws_raw = cast(Worksheet, wb.create_sheet("01_RawInput"))
    raw_headers = [
        "row_no",
        "team_id",
        "team_name",
        "judge_id",
        "innovation_input",
        "technical_input",
        "market_input",
        "demo_input",
    ]
    apply_header_style(ws_raw, raw_headers)
    for idx, row in enumerate(raw_rows, 1):
        ws_raw.append(
            [
                idx,
                row["team_id"],
                row["team_name"],
                row["judge_id"],
                row["innovation"],
                row["technical"],
                row["market"],
                row["demo"],
            ]
        )
    ws_raw.freeze_panes = "A2"
    autosize_columns(ws_raw)

    ws_calc = cast(Worksheet, wb.create_sheet("02_BackendCalc"))
    calc_headers = [
        "row_no",
        "team_id",
        "team_name",
        "judge_id",
        "innovation",
        "innovation_term(20%)",
        "technical",
        "technical_term(20%)",
        "market",
        "market_term(50%)",
        "demo",
        "demo_term(10%)",
        "weighted_from_terms",
        "weighted_stored_raw",
        "formula_delta",
        "weighted_2dp_half_up",
        "display_1dp_toFixed",
    ]
    apply_header_style(ws_calc, calc_headers)
    for idx, row in enumerate(raw_rows, 1):
        innovation_term = row["innovation"] * WEIGHTS["innovation"]
        technical_term = row["technical"] * WEIGHTS["technical"]
        market_term = row["market"] * WEIGHTS["market"]
        demo_term = row["demo"] * WEIGHTS["demo"]
        row_calc = innovation_term + technical_term + market_term + demo_term
        row_delta = abs(row_calc - row["weighted_score"])
        ws_calc.append(
            [
                idx,
                row["team_id"],
                row["team_name"],
                row["judge_id"],
                row["innovation"],
                round_half_up(innovation_term, 6),
                row["technical"],
                round_half_up(technical_term, 6),
                row["market"],
                round_half_up(market_term, 6),
                row["demo"],
                round_half_up(demo_term, 6),
                round_half_up(row_calc, 6),
                round_half_up(row["weighted_score"], 6),
                round_half_up(row_delta, 10),
                round_half_up(row_calc, 2),
                js_like_round_1dp(row_calc),
            ]
        )
    ws_calc.freeze_panes = "A2"
    autosize_columns(ws_calc)

    ws_agg = cast(Worksheet, wb.create_sheet("03_TeamAggregate"))
    agg_headers = [
        "rank_by_avg_weighted",
        "team_id",
        "team_name",
        "avg_innovation",
        "avg_technical",
        "avg_market",
        "avg_demo",
        "avg_weighted",
        "recomputed_from_avg",
        "formula_delta",
        "avg_of_row_2dp",
        "display_1dp",
        "display_collision_1dp",
        "collision_count_1dp",
    ]
    apply_header_style(ws_agg, agg_headers)
    for row in report_rows:
        ws_agg.append(
            [
                row["rank_by_avg_weighted"],
                row["team_id"],
                row["team_name"],
                round_half_up(row["avg_innovation"], 6),
                round_half_up(row["avg_technical"], 6),
                round_half_up(row["avg_market"], 6),
                round_half_up(row["avg_demo"], 6),
                round_half_up(row["avg_weighted"], 6),
                round_half_up(row["recomputed_from_avg"], 6),
                round_half_up(row["formula_delta"], 10),
                round_half_up(row["avg_of_row_2dp"], 6),
                row["ui_display_1dp"],
                row["display_collision_1dp"],
                row["collision_count_1dp"],
            ]
        )
    ws_agg.freeze_panes = "A2"
    autosize_columns(ws_agg)

    ws_rank = cast(Worksheet, wb.create_sheet("04_RankingDisputeView"))
    rank_headers = [
        "rank",
        "team_name",
        "avg_weighted(raw)",
        "display_2dp",
        "display_1dp",
        "collision_1dp",
        "dispute_note",
    ]
    apply_header_style(ws_rank, rank_headers)
    for row in report_rows:
        dispute_note = (
            "Same 1dp display with other team(s), use raw avg_weighted for strict ordering"
            if row["display_collision_1dp"]
            else "No display collision"
        )
        ws_rank.append(
            [
                row["rank_by_avg_weighted"],
                row["team_name"],
                round_half_up(row["avg_weighted"], 6),
                round_half_up(row["avg_weighted"], 2),
                row["ui_display_1dp"],
                row["display_collision_1dp"],
                dispute_note,
            ]
        )
    ws_rank.freeze_panes = "A2"
    autosize_columns(ws_rank)

    wb.save(xlsx_path)


def main() -> None:
    random.seed(42)

    repo_root = Path(__file__).resolve().parents[2]
    export_dir = repo_root / "database" / "exports"
    export_dir.mkdir(parents=True, exist_ok=True)

    raw_csv_path = export_dir / "judge_scores_realistic_mock.csv"
    report_csv_path = export_dir / "team_score_validation_report.csv"
    xlsx_path = export_dir / "final_scoring_full_pipeline_audit.xlsx"

    teams = [
        (101, "AetherMind"),
        (102, "CodeRiver"),
        (103, "NanoPulse"),
        (104, "VisionForge"),
        (105, "DataNest"),
        (106, "CloudWeave"),
        (107, "PromptPilot"),
        (108, "FlowLedger"),
    ]
    judges = [f"J{i:02d}" for i in range(1, 10)]

    # Small, realistic judge style differences.
    judge_bias = {
        "J01": (-0.4, +0.2, 0.0, -0.2),
        "J02": (+0.3, +0.1, -0.2, +0.1),
        "J03": (0.0, -0.3, +0.2, +0.1),
        "J04": (+0.1, 0.0, +0.1, -0.1),
        "J05": (-0.2, -0.1, +0.3, +0.2),
        "J06": (+0.2, +0.3, -0.1, 0.0),
        "J07": (-0.1, 0.0, 0.0, +0.3),
        "J08": (+0.4, -0.2, +0.1, -0.2),
        "J09": (0.0, +0.2, -0.2, 0.0),
    }

    # Team baseline profile.
    base = {
        101: (8.8, 8.4, 7.9, 8.3),
        102: (7.6, 8.9, 8.2, 7.8),
        103: (8.2, 7.5, 8.8, 8.0),
        104: (9.1, 8.2, 7.6, 8.9),
        105: (7.4, 7.9, 8.5, 7.2),
        106: (8.0, 8.6, 7.7, 8.1),
        107: (8.6, 7.8, 8.0, 8.7),
        108: (7.9, 8.1, 8.3, 7.9),
    }

    raw_rows: list[dict] = []
    for team_id, team_name in teams:
        for judge_id in judges:
            b = judge_bias[judge_id]
            innovation = round_to_half(
                max(0, min(10, base[team_id][0] + b[0] + random.gauss(0, 0.35)))
            )
            technical = round_to_half(
                max(0, min(10, base[team_id][1] + b[1] + random.gauss(0, 0.35)))
            )
            market = round_to_half(
                max(0, min(10, base[team_id][2] + b[2] + random.gauss(0, 0.35)))
            )
            demo = round_to_half(
                max(0, min(10, base[team_id][3] + b[3] + random.gauss(0, 0.35)))
            )
            weighted = weighted_score(innovation, technical, market, demo)

            raw_rows.append(
                {
                    "team_id": team_id,
                    "team_name": team_name,
                    "judge_id": judge_id,
                    "innovation": innovation,
                    "technical": technical,
                    "market": market,
                    "demo": demo,
                    "weighted_score": weighted,
                }
            )

    with raw_csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "team_id",
                "team_name",
                "judge_id",
                "innovation",
                "technical",
                "market",
                "demo",
                "weighted_score",
            ],
        )
        writer.writeheader()
        for row in raw_rows:
            out = dict(row)
            out["weighted_score"] = f"{row['weighted_score']:.4f}"
            writer.writerow(out)

    by_team: dict[int, list[dict]] = {}
    for row in raw_rows:
        by_team.setdefault(row["team_id"], []).append(row)

    report_rows: list[dict] = []
    formula_mismatch_count = 0
    for team_id, team_name in teams:
        rows = by_team[team_id]
        avg_innovation = sum(r["innovation"] for r in rows) / len(rows)
        avg_technical = sum(r["technical"] for r in rows) / len(rows)
        avg_market = sum(r["market"] for r in rows) / len(rows)
        avg_demo = sum(r["demo"] for r in rows) / len(rows)
        avg_weighted = sum(r["weighted_score"] for r in rows) / len(rows)

        recomputed_from_avg = weighted_score(
            avg_innovation, avg_technical, avg_market, avg_demo
        )
        delta = abs(avg_weighted - recomputed_from_avg)

        # Emulate frontend display and potential DB two-decimal storage behavior.
        ui_display_1dp = js_like_round_1dp(avg_weighted)
        avg_of_row_2dp = sum(round_half_up(r["weighted_score"], 2) for r in rows) / len(
            rows
        )

        if delta > 1e-9:
            formula_mismatch_count += 1

        report_rows.append(
            {
                "team_id": team_id,
                "team_name": team_name,
                "avg_innovation": avg_innovation,
                "avg_technical": avg_technical,
                "avg_market": avg_market,
                "avg_demo": avg_demo,
                "avg_weighted": avg_weighted,
                "recomputed_from_avg": recomputed_from_avg,
                "formula_delta": delta,
                "avg_of_row_2dp": avg_of_row_2dp,
                "ui_display_1dp": ui_display_1dp,
            }
        )

    report_rows.sort(key=lambda r: (-r["avg_weighted"], r["team_id"]))

    # Frontend uses API row order for ranking badges.
    # We track 1dp display collisions (same displayed score, different raw score).
    for idx, row in enumerate(report_rows, 1):
        row["rank_by_avg_weighted"] = idx

    rounded_buckets: dict[float, list[dict]] = {}
    for row in report_rows:
        rounded_buckets.setdefault(row["ui_display_1dp"], []).append(row)

    for row in report_rows:
        bucket = rounded_buckets[row["ui_display_1dp"]]
        row["display_collision_1dp"] = len(bucket) > 1
        row["collision_count_1dp"] = len(bucket)

    with report_csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "team_id",
                "team_name",
                "avg_innovation",
                "avg_technical",
                "avg_market",
                "avg_demo",
                "avg_weighted",
                "recomputed_from_avg",
                "formula_delta",
                "avg_of_row_2dp",
                "ui_display_1dp",
                "rank_by_avg_weighted",
                "display_collision_1dp",
                "collision_count_1dp",
            ],
        )
        writer.writeheader()
        for row in report_rows:
            out = dict(row)
            for key in [
                "avg_innovation",
                "avg_technical",
                "avg_market",
                "avg_demo",
                "avg_weighted",
                "recomputed_from_avg",
                "formula_delta",
                "avg_of_row_2dp",
            ]:
                out[key] = f"{row[key]:.6f}"
            out["ui_display_1dp"] = f"{row['ui_display_1dp']:.1f}"
            writer.writerow(out)

    export_pipeline_xlsx(
        raw_rows=raw_rows,
        report_rows=report_rows,
        formula_mismatch_count=formula_mismatch_count,
        xlsx_path=xlsx_path,
    )

    collision_rows = sum(1 for r in report_rows if r["display_collision_1dp"])

    print("Mock scoring validation completed")
    print(f"- Raw rows: {len(raw_rows)}")
    print(f"- Teams: {len(report_rows)}")
    print(f"- Formula mismatch count: {formula_mismatch_count}")
    print(f"- Teams with display collision at 1dp: {collision_rows}")
    print(f"- Raw CSV: {raw_csv_path}")
    print(f"- Report CSV: {report_csv_path}")
    print(f"- Full pipeline XLSX: {xlsx_path}")


if __name__ == "__main__":
    main()
